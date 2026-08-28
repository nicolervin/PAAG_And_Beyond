from __future__ import annotations

import sqlite3
import unittest
from contextlib import closing
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from utils import excel_io, store


class FishboneQuantityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.database_path = store.DATA_DIR / f"test_fishbone_quantities_{uuid4()}.db"
        self.database_patch = patch.object(store, "DB_PATH", self.database_path)
        self.database_patch.start()
        store.init_db()
        self.project_id = str(store.query("SELECT id FROM projects ORDER BY created_at LIMIT 1")[0]["id"])
        self.scenario_id = str(store.planning_scenarios(self.project_id)[0]["id"])
        self.section_id = store.add_assembly_section(
            self.project_id, "Decimal quantity section", "Main spine", None, ""
        )
        self.part_id = str(
            store.project_table("parts", self.project_id, "part_number").iloc[0]["id"]
        )

    def tearDown(self) -> None:
        self.database_patch.stop()
        for suffix in ("", "-wal", "-shm"):
            Path(f"{self.database_path}{suffix}").unlink(missing_ok=True)

    def assignment_rows(self) -> pd.DataFrame:
        return store.fishbone_part_assignments(self.project_id)

    def test_decimal_quantity_round_trips_through_store_and_snapshot(self) -> None:
        count = store.assign_parts_to_section(
            self.project_id,
            [self.part_id],
            self.section_id,
            quantities_by_part={self.part_id: 1.5},
        )
        self.assertEqual(count, 1)
        assignments = self.assignment_rows()
        self.assertEqual(assignments["quantity"].tolist(), [1.5])

        edited = assignments[
            ["id", "part_id", "section_id", "sequence", "quantity", "use_description", "notes"]
        ].copy()
        edited.loc[:, "quantity"] = 0.02
        store.replace_fishbone_part_assignments(self.project_id, edited)
        snapshot = store.fishbone_assignment_snapshot(self.project_id)
        store.restore_fishbone_assignment_snapshot(self.project_id, snapshot)

        self.assertEqual(self.assignment_rows()["quantity"].tolist(), [0.02])

    def test_create_part_and_assignment_accepts_decimal_quantity(self) -> None:
        part_id, assignment_id, _ = store.create_part_and_assign_to_section(
            self.project_id,
            self.section_id,
            {
                "part_number": "DECIMAL-001",
                "description": "Decimal material",
                "revision": "0",
                "model_applicability": "All",
            },
            0.125,
        )

        assignment = self.assignment_rows().loc[
            lambda rows: rows["id"].astype(str).eq(assignment_id)
        ].iloc[0]
        self.assertEqual(str(assignment["part_id"]), part_id)
        self.assertEqual(float(assignment["quantity"]), 0.125)

    def test_non_positive_and_non_finite_quantities_are_rejected(self) -> None:
        for invalid in (0, -1, float("nan"), float("inf"), "not a number"):
            with self.subTest(invalid=invalid):
                with self.assertRaisesRegex(ValueError, "greater than zero"):
                    store.assign_parts_to_section(
                        self.project_id,
                        [self.part_id],
                        self.section_id,
                        allow_additional_use=True,
                        quantities_by_part={self.part_id: invalid},
                    )

        valid = pd.DataFrame(
            [
                {
                    "id": "assignment-invalid",
                    "part_id": self.part_id,
                    "section_id": self.section_id,
                    "sequence": 10,
                    "quantity": 0,
                    "use_description": "",
                    "notes": "",
                }
            ]
        )
        with self.assertRaisesRegex(ValueError, "greater than zero"):
            store.replace_fishbone_part_assignments(self.project_id, valid)

    def test_integer_schema_is_migrated_to_real_without_losing_values(self) -> None:
        store.assign_parts_to_section(
            self.project_id,
            [self.part_id],
            self.section_id,
            quantities_by_part={self.part_id: 2.0},
        )
        with closing(sqlite3.connect(self.database_path)) as conn:
            conn.execute("PRAGMA foreign_keys = OFF")
            conn.executescript(
                """
                CREATE TABLE fishbone_part_assignments_legacy (
                    id TEXT PRIMARY KEY,
                    project_id TEXT NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                    part_id TEXT NOT NULL REFERENCES parts(id) ON DELETE CASCADE,
                    section_id TEXT NOT NULL REFERENCES assembly_sections(id),
                    sequence INTEGER NOT NULL DEFAULT 10,
                    quantity INTEGER NOT NULL DEFAULT 1,
                    use_description TEXT DEFAULT '',
                    notes TEXT DEFAULT '',
                    updated_at TEXT NOT NULL
                );
                INSERT INTO fishbone_part_assignments_legacy
                SELECT id, project_id, part_id, section_id, sequence,
                       CAST(quantity AS INTEGER), use_description, notes, updated_at
                FROM fishbone_part_assignments;
                DROP TABLE fishbone_part_assignments;
                ALTER TABLE fishbone_part_assignments_legacy
                    RENAME TO fishbone_part_assignments;
                """
            )
            conn.commit()

        store.init_db()

        with closing(sqlite3.connect(self.database_path)) as conn:
            columns = {
                row[1]: row[2] for row in conn.execute(
                    "PRAGMA table_info(fishbone_part_assignments)"
                ).fetchall()
            }
            table_sql = str(
                conn.execute(
                    "SELECT sql FROM sqlite_master WHERE type='table' "
                    "AND name='fishbone_part_assignments'"
                ).fetchone()[0]
            )
            quantity, storage_type = conn.execute(
                "SELECT quantity, typeof(quantity) FROM fishbone_part_assignments"
            ).fetchone()
        self.assertEqual(columns["quantity"].upper(), "REAL")
        self.assertIn("CHECK(quantity > 0)", table_sql)
        self.assertEqual(quantity, 2.0)
        self.assertEqual(storage_type, "real")

    def test_excel_export_keeps_decimal_fishbone_quantity(self) -> None:
        store.assign_parts_to_section(
            self.project_id,
            [self.part_id],
            self.section_id,
            quantities_by_part={self.part_id: 1.5},
        )

        workbook = load_workbook(
            BytesIO(excel_io.export_workbook(self.project_id, self.scenario_id)),
            data_only=True,
        )
        sheet = workbook["Fishbone Parts"]
        headers = [cell.value for cell in sheet[1]]
        quantity_column = headers.index("quantity") + 1
        exported_quantities = [
            sheet.cell(row=row_number, column=quantity_column).value
            for row_number in range(2, sheet.max_row + 1)
        ]
        self.assertIn(1.5, exported_quantities)


if __name__ == "__main__":
    unittest.main()
